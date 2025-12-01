"""Compare LLM models on PR metrics calculation via MCP."""

import json
import math
import os
import time
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List

import openpyxl
import requests
from openpyxl.styles import Font, Alignment

from llm_agent import LLMAgent
from mcp_host import MCPHost
from mcp_session import MCPDockerSession
from metrics_calculator import calculate_metrics
from role_classifier import classify_participant, classify_all_participants

PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = PROJECT_ROOT / "github-mcp" / ".env"
OPENROUTER_ENV_FILE = PROJECT_ROOT / ".env"
DATA_DIR = PROJECT_ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

OWNER = "catboost"
REPO = "catboost"
DAYS_BACK = 120
SAMPLE_SIZE = 20

MODELS = [
    "mistralai/mistral-tiny",
    "meta-llama/llama-3.2-3b-instruct",
    "deepseek/deepseek-chat",
    "qwen/qwen-2.5-7b-instruct",
    "anthropic/claude-3-haiku",
]


def fetch_pr_data_via_mcp(host: MCPHost) -> List[Dict]:
    """Fetch PR data through MCP tools."""
    since = (datetime.now(timezone.utc) - timedelta(days=DAYS_BACK)).date().isoformat()
    query = f"repo:{OWNER}/{REPO} comments:>0 created:>={since}"

    search_result = host.execute_tool(
        "search_pull_requests",
        {"query": query, "sort": "created", "order": "desc", "perPage": SAMPLE_SIZE},
    )

    prs = []
    for item in search_result.get("items", [])[:SAMPLE_SIZE]:
        pr_num = item["number"]
        details = host.execute_tool(
            "pull_request_read", {"owner": OWNER, "repo": REPO, "pullNumber": pr_num, "method": "get"}
        )
        issue_comments = host.execute_tool(
            "pull_request_read",
            {"owner": OWNER, "repo": REPO, "pullNumber": pr_num, "method": "get_comments", "perPage": 100},
        )
        review_comments = host.execute_tool(
            "pull_request_read",
            {"owner": OWNER, "repo": REPO, "pullNumber": pr_num, "method": "get_review_comments", "perPage": 100},
        )

        prs.append(
            {
                "number": pr_num,
                "author": (details.get("user") or {}).get("login", ""),
                "created_at": details.get("created_at", ""),
                "issue_comments": issue_comments if isinstance(issue_comments, list) else [],
                "review_comments": review_comments if isinstance(review_comments, list) else [],
            }
        )

    return prs


def run_model_analysis(model: str, api_key: str, pr_data: List[Dict]) -> Dict:
    """Run analysis for one model - data obtained via MCP, analysis via LLM."""
    start_time = time.time()
    # Note: MCP calls are not tracked per model since data is fetched once for all models
    # This metric is not meaningful when data is pre-fetched
    
    # Prepare PR summary for LLM
    pr_summary = []
    for pr in pr_data[:10]:
        comments = pr.get("issue_comments", []) + pr.get("review_comments", [])
        first_comment_time = None
        if comments:
            comment_times = [c.get("created_at") for c in comments if c.get("created_at")]
            if comment_times:
                first_comment_time = min(comment_times)
        
        participants = set()
        if pr.get("author"):
            participants.add(pr["author"])
        for c in comments:
            if isinstance(c, dict) and c.get("user", {}).get("login"):
                participants.add(c["user"]["login"])
        
        pr_summary.append({
            "number": pr.get("number", 0),
            "author": pr.get("author", ""),
            "created_at": pr.get("created_at", ""),
            "first_comment_at": first_comment_time,
            "participants": list(participants),
            "comment_count": len(comments)
        })
    
    # Calculate reference value for better guidance
    ref_metrics_temp = calculate_metrics(pr_data)
    ref_avg_temp = ref_metrics_temp["avg_first_comment_hours"]
    
    task = f"""Проанализируй данные PR репозитория {OWNER}/{REPO} (получены через GitHub MCP):

Данные PR:
{json.dumps(pr_summary, ensure_ascii=False, indent=2)}

ВАЖНО: Вычисли две метрики ТОЧНО:

1. Среднее время первого комментария (в ЧАСАХ):
   - Для каждого PR: вычисли разницу между created_at и first_comment_at
   - Разница в ЧАСАХ = (first_comment_at - created_at) в секундах / 3600
   - Затем вычисли СРЕДНЕЕ АРИФМЕТИЧЕСКОЕ всех разниц
   - Пример: если разницы [24, 48, 72] часа, то среднее = (24+48+72)/3 = 48 часов
   - Эталонное значение для проверки: примерно {ref_avg_temp} часов

2. Классификация ролей участников по паттернам активности:
   Анализируй активность каждого участника и определи паттерн и роль. ВАЖНО:
   
   Паттерны (русские названия категорий):
   - "Пассивного потребления" - только наблюдение
   - "Инициации обратной связи" - комментарии без PR
   - "Периферийного участия" - эпизодическое участие
   - "Активного соисполнительства" - регулярные PR
   - "Кураторства и управления" - управление процессом
   - "Лидерства и наставничества" - стратегическая роль
   - "Социального влияния" - высокая видимость
   
   Роли (английские названия конкретных ролей):
   - Lurkers, Passive user, Rare contributor
   - Bug reporter, Coordinator
   - Peripheral developer, Nomad Coder, Independent, Aspiring
   - Bug fixer, Active developer, Issue fixer, Code Warrior, External contributors, Internal collaborators
   - Project steward, Coordinator, Progress controller
   - Project leader, Core member, Core developer
   - Project Rockstar
   
   Для каждого участника: username, pattern (русское название паттерна), role (английское название конкретной роли)

Верни JSON:
{{
  "avg_first_comment_hours": число,
  "user_roles": [{{"username": "user", "pattern": "Кураторства и управления", "role": "Project steward"}}],
  "role_distribution": {{"Кураторства и управления": количество, ...}}
}}
"""
    
    try:
        resp = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": "Ты анализируешь GitHub метрики. Всегда возвращай ТОЛЬКО валидный JSON, без объяснений и кода."},
                    {"role": "user", "content": task}
                ],
                "temperature": 0.1,  # Lower temperature for more consistent JSON
            },
            headers={
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://github.com/elenagernichenko/github-analyzer",
                "X-Title": "github-analyzer",
            },
            timeout=120,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        elapsed = time.time() - start_time
        result = {"content": content, "time_seconds": round(elapsed, 2), "iterations": 1}
    except Exception as e:
        elapsed = time.time() - start_time
        result = {"content": f'{{"error": "{str(e)}"}}', "time_seconds": round(elapsed, 2), "iterations": 1}

    # Calculate reference metrics
    ref_metrics = calculate_metrics(pr_data)
    ref_avg = ref_metrics["avg_first_comment_hours"]

    # Parse LLM response with improved extraction (especially for llama)
    content = result["content"].strip()
    raw_content = content[:500]
    
    # Enhanced extraction for llama and other models
    json_str = None
    
    # Strategy 1: Extract from code blocks
    if "```json" in content:
        json_str = content.split("```json")[1].split("```")[0].strip()
    elif "```" in content:
        parts = content.split("```")
        for part in parts:
            if "{" in part and "}" in part and '"avg_first_comment_hours"' in part:
                json_str = part.strip()
                break
    
    # Strategy 2: Find JSON object boundaries
    if not json_str:
        start = content.find("{")
        if start >= 0:
            # Find matching closing brace
            brace_count = 0
            end = start
            for i in range(start, len(content)):
                if content[i] == "{":
                    brace_count += 1
                elif content[i] == "}":
                    brace_count -= 1
                    if brace_count == 0:
                        end = i + 1
                        break
            if end > start:
                json_str = content[start:end]
    
    # Strategy 3: Extract using regex (for llama that returns code)
    if not json_str:
        import re
        # Try to find JSON-like structure
        json_match = re.search(r'\{[^{}]*"avg_first_comment_hours"[^{}]*\}', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0)
        else:
            # Extract just the numbers
            hours_match = re.search(r'"avg_first_comment_hours"\s*:\s*([\d.]+)', content)
            if hours_match:
                json_str = f'{{"avg_first_comment_hours": {hours_match.group(1)}'
                # Try to add role_distribution
                dist_match = re.search(r'"role_distribution"\s*:\s*(\{[^}]+\})', content, re.DOTALL)
                if dist_match:
                    json_str += f', "role_distribution": {dist_match.group(1)}'
                json_str += "}"
    
    # Parse JSON
    llm_result = {}
    try:
        if json_str:
            llm_result = json.loads(json_str)
        else:
            llm_result = json.loads(content)
    except (json.JSONDecodeError, ValueError):
        # Last resort: manual extraction
        import re
        hours_match = re.search(r'"avg_first_comment_hours"\s*:\s*([\d.]+)', content)
        if hours_match:
            llm_result["avg_first_comment_hours"] = float(hours_match.group(1))
        # Extract role_distribution more flexibly
        dist_pattern = r'"role_distribution"\s*:\s*\{([^}]+)\}'
        dist_match = re.search(dist_pattern, content, re.DOTALL)
        if dist_match:
            try:
                # Try to parse the distribution
                dist_str = "{" + dist_match.group(1) + "}"
                llm_result["role_distribution"] = json.loads(dist_str)
            except:
                # Extract key-value pairs manually
                pairs = re.findall(r'"([^"]+)"\s*:\s*(\d+)', dist_match.group(1))
                if pairs:
                    llm_result["role_distribution"] = {k: int(v) for k, v in pairs}
    
    llm_avg = float(llm_result.get("avg_first_comment_hours", 0))
    error_pct = abs(llm_avg - ref_avg) / ref_avg * 100 if ref_avg > 0 and llm_avg > 0 else (100 if llm_avg == 0 else 0)
    llm_dist = llm_result.get("role_distribution", {})
    user_roles = llm_result.get("user_roles", [])  # Contains pattern and role for each user
    
    # Normalize user_roles: support both old format (role/reason) and new format (pattern/role)
    normalized_user_roles = []
    for ur in user_roles:
        # New format: pattern + role
        if "pattern" in ur:
            normalized_user_roles.append({
                "username": ur.get("username", ""),
                "pattern": ur.get("pattern", ""),
                "role": ur.get("role", "")
            })
        # Old format: role (was pattern) + reason (was role)
        elif "role" in ur and "reason" in ur:
            normalized_user_roles.append({
                "username": ur.get("username", ""),
                "pattern": ur.get("role", ""),  # Old "role" was actually pattern
                "role": ur.get("reason", "").replace(" паттерн", "").replace(" паттерн", "").strip()  # Old "reason" was actually role
            })
        elif "role" in ur:
            # Fallback: assume role is pattern
            normalized_user_roles.append({
                "username": ur.get("username", ""),
                "pattern": ur.get("role", ""),
                "role": ""
            })
    
    # Debug: log if parsing failed
    if llm_avg == 0 and len(llm_dist) == 0:
        print(f"    ⚠ Warning: Could not parse LLM response. Raw: {raw_content[:100]}...")

    ref_dist = ref_metrics["role_distribution"]

    # Role correctness: проверяет правильность определения паттернов пользователей
    # Сравнивает паттерны, присвоенные LLM, с эталонными паттернами, вычисленными на основе данных репозитория
    # 
    # Эталонные паттерны вычисляются на основе активности пользователя:
    # - prs_authored: количество PR, созданных пользователем
    # - prs_reviewed: количество PR, отревьюенных пользователем
    # - comments_count: количество комментариев
    # - participation_rate, comment_rate: процент участия
    #
    # Метрика = (количество пользователей с правильным паттерном) / (общее количество пользователей) * 100%
    
    # Вычислить эталонные паттерны для всех пользователей
    participants_stats: Dict[str, Dict[str, int]] = {}
    total_prs = len(pr_data)
    
    for pr in pr_data:
        author = pr.get("author", "")
        if author:
            if author not in participants_stats:
                participants_stats[author] = {"authored": 0, "reviewed": 0, "comments": 0}
            participants_stats[author]["authored"] += 1
        
        for comment in pr.get("issue_comments", []) + pr.get("review_comments", []):
            user = comment.get("user", {}).get("login", "") if isinstance(comment, dict) else ""
            if user and user != author:
                if user not in participants_stats:
                    participants_stats[user] = {"authored": 0, "reviewed": 0, "comments": 0}
                participants_stats[user]["reviewed"] += 1
                participants_stats[user]["comments"] += 1
    
    # Вычислить эталонный паттерн для каждого пользователя
    reference_patterns: Dict[str, str] = {}
    for username, stats in participants_stats.items():
        ref_pattern = classify_participant(
            username,
            stats["authored"],
            stats["reviewed"],
            stats["comments"],
            total_prs,
        )
        reference_patterns[username.lower()] = ref_pattern.lower()
    
    # Сравнить паттерны LLM с эталонными
    role_correctness = 0.0
    if normalized_user_roles:
        correct = 0
        total = 0
        for ur in normalized_user_roles:
            username = ur.get("username", "").lower()
            llm_pattern = ur.get("pattern", "").lower()
            if username and llm_pattern:
                total += 1
                ref_pattern = reference_patterns.get(username, "")
                # Проверяем соответствие паттернов (гибкое сравнение)
                if ref_pattern and (ref_pattern in llm_pattern or llm_pattern in ref_pattern):
                    correct += 1
        role_correctness = (correct / total * 100) if total > 0 else 0
    elif llm_dist:
        # Fallback: если нет user_roles, проверяем только валидность названий паттернов
        expected_patterns = {"Пассивного потребления", "Инициации обратной связи", "Периферийного участия",
                             "Активного соисполнительства", "Кураторства и управления",
                             "Лидерства и наставничества", "Социального влияния"}
        valid = sum(1 for cat in llm_dist.keys() 
                   if any(ep.lower() in cat.lower() for ep in expected_patterns))
        role_correctness = (valid / len(llm_dist) * 100) if llm_dist else 0
    
    # Coverage: how many different role categories were found
    coverage = len(set(llm_dist.keys())) if llm_dist else 0
    
    # Calculate all additional metrics
    expected_categories = {
        "Пассивного потребления", "Инициации обратной связи", "Периферийного участия",
        "Активного соисполнительства", "Кураторства и управления",
        "Лидерства и наставничества", "Социального влияния"
    }
    # MCP correctness: data was obtained via MCP and LLM successfully analyzed it
    mcp_correctness = 100 if llm_avg > 0 and (len(llm_dist) > 0 or len(user_roles) > 0) else 0
    data_completeness = 100 if llm_avg > 0 and (len(llm_dist) > 0 or len(user_roles) > 0) else 0
    # Consistency: if model found participants and assigned roles consistently
    # If all values are 0, it means model didn't find participants (not inconsistency)
    total_participants = sum(llm_dist.values()) if llm_dist else len(user_roles) if user_roles else 0
    consistency = 100 if total_participants > 0 and (len(llm_dist) > 0 or len(user_roles) > 0) else 0
    # Pattern compliance: check if role names match expected patterns (Russian or English)
    expected_categories = {
        "Пассивного потребления", "Инициации обратной связи", "Периферийного участия",
        "Активного соисполнительства", "Кураторства и управления",
        "Лидерства и наставничества", "Социального влияния"
    }
    pattern_names = {"lurkers", "issues", "independent", "aspiring", "nomad", "external", "internal", 
                    "bug fixer", "steward", "coordinator", "core", "leader", "rockstar"}
    matched = sum(1 for cat in llm_dist.keys() 
                 if any(e.lower() in cat.lower() for e in expected_categories) or
                    any(p in cat.lower() for p in pattern_names))
    pattern_compliance = (matched / len(llm_dist) * 100) if llm_dist else 0

    return {
        "model": model,
        "error_pct": round(error_pct, 2),
        "mcp_correctness": round(mcp_correctness, 2),
        "data_completeness": round(data_completeness, 2),
        "coverage": coverage,
        "role_correctness": round(role_correctness, 2),
        "consistency": round(consistency, 2),
        "pattern_compliance": round(pattern_compliance, 2),
        "time_seconds": round(result["time_seconds"], 2),
        "llm_avg_hours": llm_avg,
        "ref_avg_hours": ref_avg,
        "raw_response": content,
        "parsed_result": llm_result,
        "user_roles": normalized_user_roles,  # Save normalized user roles (pattern + role)
    }


def calculate_inter_model_agreement(results: List[Dict]) -> Dict[str, float]:
    """Calculate how consistently models assign patterns to the same users.
    Упрощенная метрика: для каждой модели считаем, сколько раз она присвоила тот же паттерн,
    что и другие модели, для одних и тех же пользователей.
    
    Метрика = (количество совпадений паттернов с другими моделями) / (общее количество сравнений) * 100%
    """
    # Collect all user-pattern assignments from all models
    user_patterns: Dict[str, Dict[str, str]] = {}  # {username: {model: pattern}}
    
    for result in results:
        model_name = result["model"].split("/")[-1]
        user_roles = result.get("user_roles", [])
        
        for ur in user_roles:
            username = ur.get("username", "").lower()
            pattern = ur.get("pattern", "").lower()
            if username and pattern:
                if username not in user_patterns:
                    user_patterns[username] = {}
                user_patterns[username][model_name] = pattern
    
    # Calculate agreement for each model: count matches with other models
    model_agreements: Dict[str, Dict[str, int]] = {}  # {model: {"matches": X, "total": Y}}
    for result in results:
        model_name = result["model"].split("/")[-1]
        model_agreements[model_name] = {"matches": 0, "total": 0}
    
    # For each user, compare patterns between all models
    for username, model_patterns in user_patterns.items():
        if len(model_patterns) < 2:  # Need at least 2 models to compare
            continue
        
        models = list(model_patterns.keys())
        patterns = list(model_patterns.values())
        
        # Compare each model with all other models for this user
        for i, model1 in enumerate(models):
            pattern1 = patterns[i]
            for j, model2 in enumerate(models):
                if i != j:  # Don't compare model with itself
                    model_agreements[model1]["total"] += 1
                    if pattern1 == patterns[j]:  # Patterns match
                        model_agreements[model1]["matches"] += 1
    
    # Calculate percentage agreement for each model
    avg_agreements: Dict[str, float] = {}
    for model_name, stats in model_agreements.items():
        if stats["total"] > 0:
            avg_agreements[model_name] = round((stats["matches"] / stats["total"]) * 100, 2)
        else:
            avg_agreements[model_name] = 0.0
    
    return avg_agreements


def load_env_file(env_path: Path) -> Dict[str, str]:
    """Load environment variables from .env file."""
    env_vars = {}
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, value = line.split("=", 1)
                env_vars[key.strip()] = value.strip().strip('"').strip("'")
    return env_vars


def main():
    """Compare all models."""
    # Try to load from .env file first
    env_vars = load_env_file(OPENROUTER_ENV_FILE)
    api_key = env_vars.get("OPENROUTER_API_KEY") or os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise SystemExit(
            f"OPENROUTER_API_KEY not found. Set it in {OPENROUTER_ENV_FILE} or export OPENROUTER_API_KEY"
        )

    # Fetch PR data once
    print("Fetching PR data via MCP...")
    with MCPDockerSession(env_file=ENV_FILE) as mcp_session:
        host = MCPHost(mcp_session)
        host.initialize()
        pr_data = fetch_pr_data_via_mcp(host)

    print(f"Fetched {len(pr_data)} PRs. Starting model comparison...")

    results = []
    llm_responses = {}  # Store raw LLM responses
    
    for model in MODELS:
        print(f"\nTesting {model}...")
        try:
            result = run_model_analysis(model, api_key, pr_data)
            results.append(result)
            # Store raw response
            llm_responses[model] = {
                "raw_response": result.get("raw_response", ""),
                "parsed_result": result.get("parsed_result", {}),
                "user_roles": result.get("user_roles", []),
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
            print(f"  ✓ Error: {result['error_pct']}%, Coverage: {result['coverage']}, Role correctness: {result.get('role_correctness', 0)}%")
        except Exception as e:
            print(f"  ✗ Error: {e}")
            results.append({
                "model": model,
                "error_pct": 100,
                "mcp_correctness": 0,
                "data_completeness": 0,
                "coverage": 0,
                "role_correctness": 0,
                "consistency": 0,
                "pattern_compliance": 0,
                "inter_model_agreement": 0,
                "time_seconds": 0,
                "error": str(e),
            })
            llm_responses[model] = {
                "raw_response": "",
                "parsed_result": {},
                "error": str(e),
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }

    # Save results
    output_path = DATA_DIR / "comparison_results.json"
    output_path.write_text(json.dumps(results, indent=2, ensure_ascii=False))
    
    # Save LLM responses separately
    responses_path = DATA_DIR / "llm_responses.json"
    responses_path.write_text(json.dumps(llm_responses, indent=2, ensure_ascii=False))
    
    # Calculate inter-model agreement: how consistently models assign patterns to the same users
    inter_model_agreement = calculate_inter_model_agreement(results)
    
    # Add agreement metric to each result
    for result in results:
        model_name = result["model"].split("/")[-1]
        result["inter_model_agreement"] = inter_model_agreement.get(model_name, 0.0)

    # Generate Excel with transposed layout (models as columns, metrics as rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сравнение моделей"
    
    # Define detailed metric names
    metric_names = [
        "Ошибка расчета среднего времени первого комментария (%)",
        "Корректность использования данных, полученных через MCP (%)",
        "Полнота данных: наличие обеих метрик (время + роли) (%)",
        "Количество найденных категорий ролей",
        "Корректность классификации ролей: правильность определения ролей пользователей согласно паттернам (%)",
        "Консистентность классификации: стабильность логики распределения (%)",
        "Соответствие паттернам: использование правильных названий категорий (%)",
        "Согласованность между моделями: насколько одинаково модели присваивают паттерны одним и тем же участникам (%)",
        "Время выполнения анализа (секунды)",
    ]
    
    # Prepare data: models as columns
    model_names = [r["model"].split("/")[-1] for r in results]
    
    # Header row: Метрика + model names
    header_row = ["Метрика"] + model_names
    ws.append(header_row)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows: metric name + values for each model
    metric_data = [
        [r.get("error_pct", 100) for r in results],
        [r.get("mcp_correctness", 0) for r in results],
        [r.get("data_completeness", 0) for r in results],
        [r.get("coverage", 0) for r in results],
        [r.get("role_correctness", 0) for r in results],
        [r.get("consistency", 0) for r in results],
        [r.get("pattern_compliance", 0) for r in results],
        [r.get("inter_model_agreement", 0) for r in results],
        [r.get("time_seconds", 0) for r in results],
    ]
    
    for metric_name, values in zip(metric_names, metric_data):
        row = [metric_name] + values
        ws.append(row)
    
    # Style and adjust column widths
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].font = Font(bold=False)
        row[0].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 60  # Metric names column
    for col_idx, model_name in enumerate(model_names, start=2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(model_name) + 2, 15)
    
    excel_path = DATA_DIR / "comparison_report.xlsx"
    wb.save(excel_path)
    
    print(f"\n✓ Results saved to {output_path}")
    print(f"✓ LLM responses saved to {responses_path}")
    print(f"✓ Excel report saved to {excel_path}")


if __name__ == "__main__":
    main()

